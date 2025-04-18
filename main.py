import csv

def read_excel_manual(filename):
    # Membaca file excel sederhana dalam bentuk CSV (tanpa menggunakan pandas atau openpyxl)
    import openpyxl
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append({
            "id": row[0],
            "service": row[1],
            "price": row[2]
        })
    return data

# Membership Functions
def triangular(x, a, b, c):
    if x <= a or x >= c:
        return 0
    elif a < x < b:
        return (x - a) / (b - a)
    elif b <= x < c:
        return (c - x) / (c - b)
    return 0

def fuzzify_service(x):
    return {
        "low": triangular(x, 0, 30, 50),
        "medium": triangular(x, 30, 50, 70),
        "high": triangular(x, 50, 70, 100)
    }

def fuzzify_price(x):
    return {
        "cheap": triangular(x, 25000, 30000, 40000),
        "medium": triangular(x, 30000, 40000, 50000),
        "expensive": triangular(x, 40000, 50000, 55000)
    }

# Nilai skor (output defuzzification centroid dari setiap rule)
score_values = {
    "very_bad": 10,
    "bad": 30,
    "fair": 50,
    "good": 70,
    "very_good": 90
}

# Rule base Fuzzy Inference
def inference(service, price):
    rules = []

    # Contoh aturan:
    rules.append((min(service["high"], price["cheap"]), "very_good"))
    rules.append((min(service["high"], price["medium"]), "good"))
    rules.append((min(service["high"], price["expensive"]), "fair"))
    rules.append((min(service["medium"], price["cheap"]), "good"))
    rules.append((min(service["medium"], price["medium"]), "fair"))
    rules.append((min(service["medium"], price["expensive"]), "bad"))
    rules.append((min(service["low"], price["cheap"]), "fair"))
    rules.append((min(service["low"], price["medium"]), "bad"))
    rules.append((min(service["low"], price["expensive"]), "very_bad"))

    # Agregasi (Defuzzifikasi menggunakan metode centroid)
    numerator = 0
    denominator = 0
    for weight, label in rules:
        val = score_values[label]
        numerator += weight * val
        denominator += weight

    return numerator / denominator if denominator != 0 else 0

def write_to_excel(data, filename):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["ID Restoran", "Service Quality", "Price", "Score"])
    for d in data:
        ws.append([d["id"], d["service"], d["price"], round(d["score"], 2)])
    wb.save(filename)

def main():
    data = read_excel_manual("restoran.xlsx")
    scored_data = []

    for item in data:
        service_membership = fuzzify_service(item["service"])
        price_membership = fuzzify_price(item["price"])
        score = inference(service_membership, price_membership)
        item["score"] = score
        scored_data.append(item)

    # Urutkan berdasarkan skor tertinggi
    top_5 = sorted(scored_data, key=lambda x: x["score"], reverse=True)[:5]

    # Tulis ke Excel
    write_to_excel(top_5, "peringkat.xlsx")

    # Tampilkan
    print("Top 5 Restoran Terbaik:")
    for r in top_5:
        print(f"ID: {r['id']}, Service: {r['service']}, Price: {r['price']}, Score: {round(r['score'], 2)}")

if __name__ == "__main__":
    main()
